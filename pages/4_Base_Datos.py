from pickle import NONE
import streamlit as st
import pandas as pd
import os
import io
import time
import zipfile
import numpy as np
import matplotlib.pyplot as plt
import tempfile
from utils.authentication import Authentication
from utils.dataframe_utils import custom_dataframe_explorer
from utils.config import get_config
from utils.download_utils import GestorDescargas, preparar_ruta_destino, CombinadorPDF, sanitizar_nombre_archivo

# --- Verificar si los datos están completamente cargados ---
if not st.session_state.get("data_fully_loaded", False):
    st.warning("Los datos aún se están cargando. Por favor, espera en la página principal hasta que se complete la carga.")
    st.stop()

# Importar funciones centralizadas para acceso a datos
from pages.utils_3 import get_data_loader_instance, get_column_mapping

# Load custom CSS
def load_css():
    with open(os.path.join("assets", "styles.css")) as f:
        st.markdown(f"<style>{f.read()}</style>", unsafe_allow_html=True)

try:
    load_css()
except Exception as e:
    st.session_state.toast_message = f"No se pudo cargar el CSS personalizado: {e}"
    st.session_state.toast_icon = '🚨'  # Icono de error (sirena)

# Initialize authentication
authentication = Authentication()
# Main page content
st.title(":file_cabinet: Base de Datos Portal de Proveedores")
st.markdown("Impulsado por PalmaTerra 360 | Módulo Facturas :speech_balloon: :llama:", help="Explora y consulta la base de datos de la empresa")

# Importar las funciones centralizadas desde el módulo de utilidades
from utils.chatbot_supabase import init_chatbot_supabase_client, get_chatbot_filter_options, get_filtered_data_multiselect

# Inicializar variables para el sistema de toasts
if 'toast_message' not in st.session_state:
    st.session_state.toast_message = None
if 'toast_icon' not in st.session_state:
    st.session_state.toast_icon = None

# Inicializar variables para el sistema de toasts en sidebar
if 'sidebar_toast_message' not in st.session_state:
    st.session_state.sidebar_toast_message = None
if 'sidebar_toast_icon' not in st.session_state:
    st.session_state.sidebar_toast_icon = None

# Inicializar variables para las selecciones guardadas
if 'saved_selections' not in st.session_state:
    st.session_state.saved_selections = pd.DataFrame()
if 'saved_selections_desglosado' not in st.session_state:
    st.session_state.saved_selections_desglosado = pd.DataFrame()

# Usar las funciones centralizadas con caché
supabase_client_chatbot = init_chatbot_supabase_client()

if supabase_client_chatbot:
    chatbot_filter_opts = get_chatbot_filter_options(supabase_client_chatbot)
else:
    # Fallback to empty options if Supabase client failed
    chatbot_filter_opts = {key: [] for key in ['obras', 'proveedores', 'subcategorias', 'categorias']}
    st.session_state.toast_message = "No se pudo conectar a Supabase. Los filtros no estarán disponibles."
    st.session_state.toast_icon = '🚨'  # Icono de error (sirena)

# Mostrar toasts si hay mensajes pendientes
if st.session_state.toast_message:
    st.toast(st.session_state.toast_message, icon=st.session_state.toast_icon)
    st.session_state.toast_message = None
    st.session_state.toast_icon = None

# Mostrar toasts en la sidebar (aunque técnicamente los toasts son globales)
if st.session_state.sidebar_toast_message:
    st.toast(st.session_state.sidebar_toast_message, icon=st.session_state.sidebar_toast_icon)
    st.session_state.sidebar_toast_message = None
    st.session_state.sidebar_toast_icon = None

# # CSS para quitar borde del formulario
# st.markdown('''
# <style>
#     .stForm [data-testid="stForm"] {
#         border: none;
#         padding-top: 0;
#         padding-bottom: 0;
#     }
# </style>
# ''', unsafe_allow_html=True)

# Formulario de filtros para consulta de facturas dentro de un expander
with st.expander("🔍 Búsqueda", expanded=True):
    # Importar datetime al principio, fuera de los bloques anidados
    import datetime
    today = datetime.date.today()
    jan_1 = datetime.date(2018, 1, 1)
    dec_31 = datetime.date(today.year, 12, 31)
    
    # Definir el formulario
    with st.form(key="consulta_facturas_form",border=False):
        col1, col2, col3 = st.columns(3)
        
        with col1:
            # Filtro de Obras
            obras_seleccionadas = st.multiselect(
                "Obra:",
                options=chatbot_filter_opts.get('obras', []),
                placeholder="Seleccione una obra..."
            )
        
        with col2:
            # Filtro de Proveedores
            proveedores_seleccionados = st.multiselect(
                "Proveedor:",
                options=chatbot_filter_opts.get('proveedores', []),
                placeholder="Filtrar por proveedor..."
            )
        
        with col3:
            col3_1, col3_2 = st.columns(2)
            # Filtro de Fecha con rango (sin valor por defecto)
            try:
                with col3_1:
                    fecha_inicio = st.date_input(
                        "Fecha Inicio:",
                        None,  # Valor inicial vacío
                        min_value=jan_1,
                        max_value=today,
                        format="DD/MM/YYYY",
                    help="Selecciona la fecha de inicio"
                )
                with col3_2:
                    fecha_fin = st.date_input(
                        "Fecha Fin:",
                        None,  # Valor inicial vacío
                        min_value=jan_1,
                        max_value=today,
                        format="DD/MM/YYYY",
                    help="Selecciona la fecha de fin"
                )
            except Exception:
                # Fallback si hay un problema con None
                fecha_inicio = None
                fecha_fin = None

        estatus_date = st.columns([1, 1])
        with estatus_date[0]:
            estatus_seleccionados = st.segmented_control(
                "Estatus:",
                options=['Pagada', 'Proceso de Pago', 'RevisaRes'],
                selection_mode="multi",
            )
        with estatus_date[1]:
            fecha_seleccionada = st.segmented_control(
                "Fecha:",
                options=['Fecha Factura', 'Fecha Recepción', 'Fecha Pagado', 'Fecha Autorización'],
                selection_mode="single",
            )
        


        # Botones centrados en dos columnas
        col_buttons = st.columns([1, 1])
        
        with col_buttons[0]:
            buscar_button = st.form_submit_button(
                "🔍 BUSCAR",
                use_container_width=True,
                type="primary"
            )
        
        with col_buttons[1]:
            limpiar_button = st.form_submit_button(
                "🧹 LIMPIAR",
                use_container_width=True
            )


# --- Carga de Datos Principal desde el cargador centralizado mejorado ---
try:
    # Importar la función que creará el dataframe de contabilidad
    
    
    # Inicializar o recuperar data desde session_state si existe
    if 'saved_data' not in st.session_state:
        st.session_state.saved_data = pd.DataFrame()
    if 'saved_data_contabilidad' not in st.session_state:
        st.session_state.saved_data_contabilidad = pd.DataFrame()
        
    # Usar datos guardados o inicializar nuevos
    data = st.session_state.saved_data
    # data_contabilidad will now also be fetched and stored in session_state if needed, or fetched fresh
    data_contabilidad = st.session_state.get('saved_data_contabilidad', pd.DataFrame())
    
    # Manejar el botón de limpiar filtros
    if limpiar_button:
        # Limpiar explorers de dataframe si existen
        if 'desglosado_explorer' in st.session_state:
            st.session_state.pop('desglosado_explorer')
        if 'concentrado_explorer' in st.session_state:
            st.session_state.pop('concentrado_explorer')
        # Reiniciar los dataframes en session_state
        st.session_state.saved_data = pd.DataFrame()
        st.session_state.saved_data_contabilidad = pd.DataFrame()
        data = pd.DataFrame()
        data_contabilidad = pd.DataFrame()
        # Recargar la página para restablecer todos los widgets
        st.rerun()
    
    # Solo realizar la búsqueda cuando se presiona el botón BUSCAR
    if buscar_button:
        # Limpiar mensajes anteriores para asegurar actualización
        st.session_state.toast_message = None
        st.session_state.toast_icon = None
        st.session_state.sidebar_toast_message = None
        st.session_state.sidebar_toast_icon = None
        
        # Limpiar los filtros de segundo nivel pero conservar las columnas seleccionadas
        explorers = ['desglosado_explorer', 'concentrado_explorer']
        for explorer_id in explorers:
            if explorer_id in st.session_state:
                # Guardar las columnas seleccionadas (primer nivel)
                selected_columns = st.session_state[explorer_id].get('_columns_to_filter_selection', [])
                
                # Crear un nuevo diccionario solo con las columnas seleccionadas
                # Esto eliminará todos los filtros de segundo nivel
                st.session_state[explorer_id] = {'_columns_to_filter_selection': selected_columns}
        # Convertir las obras seleccionadas a sus cuentas_gasto correspondientes
        # selected_cuentas_gasto = []
        # for obra in obras_seleccionadas:
        #     if obra in chatbot_filter_opts['obra_to_cuenta_gasto']:
        #         cuenta_gasto = chatbot_filter_opts['obra_to_cuenta_gasto'][obra]
        #         selected_cuentas_gasto.append(str(cuenta_gasto))
                
        # Definir columnas para portal_desglosado
        desglosado_columns = (
            "obra, tipo_gasto, cuenta_gasto, proveedor, residente, folio, estatus, "
            "fecha_factura, fecha_recepcion, fecha_pagada, fecha_autorizacion, clave_producto, clave_unidad, "
            "categoria_id, subcategoria, descripcion, cantidad, unidad, precio_unitario, subtotal, descuento, venta_tasa_0, venta_tasa_16, moneda, total_iva, "
            "total_ish, retencion_isr, retencion_iva, total, serie, url_pdf, url_oc, url_rem, xml_uuid, sat"
        )
        
        # Definir columnas para portal_contabilidad
        contabilidad_columns = (
            "obra, tipo_gasto, cuenta_gasto, proveedor, residente, folio, estatus, "
            "fecha_factura, fecha_recepcion, fecha_pagada, fecha_autorizacion, subtotal, descuento, venta_tasa_0, venta_tasa_16, moneda, total_iva, "
            "total_ish, retencion_isr, retencion_iva, total, serie, url_pdf, url_oc, url_rem, xml_uuid"
        )

        # Obtener datos para la tabla 'portal_desglosado'
        data = get_filtered_data_multiselect(
            _client=supabase_client_chatbot,
            table_name="portal_desglosado",
            select_columns=desglosado_columns,
            obras_seleccionadas=obras_seleccionadas,
            proveedores_seleccionados=proveedores_seleccionados,
            fecha_inicio=fecha_inicio,
            fecha_fin=fecha_fin,
            estatus_seleccionados=estatus_seleccionados,
            fecha_seleccionada=fecha_seleccionada
        )
        
        # Obtener datos para la tabla 'portal_contabilidad'
        data_contabilidad = get_filtered_data_multiselect(
            _client=supabase_client_chatbot,
            table_name="portal_contabilidad", 
            select_columns=contabilidad_columns,  # O especificar columnas si se conocen
            obras_seleccionadas=obras_seleccionadas, 
            proveedores_seleccionados=proveedores_seleccionados,
            fecha_inicio=fecha_inicio,
            fecha_fin=fecha_fin,
            estatus_seleccionados=estatus_seleccionados,
            fecha_seleccionada=fecha_seleccionada
        )

        # Guardar en session_state para persistencia entre reruns
        st.session_state.saved_data = data
        st.session_state.saved_data_contabilidad = data_contabilidad
        
        # Mostrar un mensaje de éxito si se encontraron datos
        if not data.empty:
            st.session_state.toast_message = f'Se encontraron **{len(data_contabilidad)} facturas** que coinciden con los filtros seleccionados.'
            st.session_state.toast_icon = '✅'  # Icono de éxito (check verde)
            # Forzar rerun para actualizar y mostrar toast
            st.rerun()

    # Sólo mostrar advertencia cuando sea una búsqueda activa (después del clic en buscar)
    if buscar_button and (data is None or data.empty):
        st.session_state.toast_message = "No se encontraron datos que coincidan con los filtros seleccionados."
        st.session_state.toast_icon = '⚠️'  # Icono de advertencia
        # Forzar rerun para actualizar y mostrar toast
        st.rerun()
        # Configurar valores por defecto para evitar errores posteriores
        data = pd.DataFrame()
        column_mapping = {}
    else:
        # Obtener el mapeo de columnas desde la configuración centralizada
        column_mapping = get_column_mapping()

except Exception as e:
    st.session_state.toast_message = f"Ocurrió un error inesperado durante la obtención de datos: {str(e)}"
    st.session_state.toast_icon = '🚨'  # Icono de error (sirena)
    # Fallback seguro
    data = pd.DataFrame()
    column_mapping = {}

# Contenido principal - Sección de Datos con pestañas
# Display tabs if either data or data_contabilidad has content
if not data.empty or not data_contabilidad.empty:
    # Usar los datos originales (sin filtros del sidebar) para el dataframe
    # Listas de columnas para formateo
    currency_columns = ["Subtotal", "Precio Unitario", "Total", "IVA 16%", "Descuento", "Retención IVA", "Retención ISR", "ISH", 'Venta Tasa 0%', 'Venta Tasa 16%']
    URL_columns = ["Factura", "Orden de Compra", "Remisión"] # Asumiendo que esta columna existe y no fue renombrada

    # Renombrar ANTES de pasar a dataframe_explorer
    display_data_renamed = data.copy()
    display_data_renamed.rename(columns=column_mapping, inplace=True, errors='ignore') # Ignorar si alguna columna del map no existe
    
    # Renombrar datos en session_state para uso posterior en dialogs
    if not st.session_state.saved_data_contabilidad.empty:
        st.session_state.saved_data_contabilidad.rename(columns=column_mapping, inplace=True, errors='ignore')
    if not st.session_state.saved_data.empty:
        st.session_state.saved_data.rename(columns=column_mapping, inplace=True, errors='ignore')
        
    # Guardar dataframes renombrados en session_state para facilitar acceso desde dialogs
    st.session_state.df_desglosado = display_data_renamed.copy()
    st.session_state.df_concentrado = st.session_state.saved_data_contabilidad.copy() if not st.session_state.saved_data_contabilidad.empty else pd.DataFrame()
    
    # Initialize filtered_df_renamed before tabs, it will be updated by the explorer in tab1
    filtered_df_renamed = display_data_renamed.copy()

    # Crear pestañas para visualización de datos
    tab1, tab2 = st.tabs(["Concentrado", "Desglosado"])

    # Pestaña 1 - Vista Desglosada (original)
    with tab2:
        # Definir explicitamente el tipo de cada columna para el filtrado

        # Columnas para selección múltiple (categorías/enumeraciones)
        multiselect_filter_cols = [
            "Obra", "Subcategoría", "Categoría", "Proveedor", "Residente", "Tipo Gasto",
            "Estatus", "Unidad", "Moneda", "Factura", "Orden de Compra", "Remisión", "Venta Tasa 0%", "Descuento"
        ]

        # Columnas de fecha para filtrado con calendario
        fecha_columns = [
            'Fecha Factura', 'Fecha Recepción', 'Fecha Pagado', 'Fecha Autorización'
        ]

        # Columnas numéricas para filtrado con slider
        numeric_columns = [
            "Subtotal", "Precio Unitario", "Total", "IVA 16%", 
            "Retención IVA", "Retención ISR", "ISH", 'Venta Tasa 16%',
            "Cantidad","Venta Tasa 0%", "Descuento"
        ]

        # Columnas de texto para búsqueda por patrones
        text_columns = [
            "Concepto", "Cuenta Gasto", "Clave Producto", "Clave Unidad", "UUID", "Descripción", "Serie", "sat"
        ]

        # Manejar potenciales errores con el custom_dataframe_explorer
        try:
            # Si data está vacío, evitar ejecutar el explorer
            if display_data_renamed.empty:
                # filtered_df_renamed is already display_data_renamed.copy() via pre-tab initialization
                pass 
            else:
                # Definir las columnas que no deben aparecer en las opciones de filtrado
                excluded_columns = numeric_columns + ["Factura", "Orden de Compra", "Remisión", "Cuenta Gasto", "sat", "Serie"]
                
                filtered_df_renamed = custom_dataframe_explorer(
                    df=display_data_renamed, 
                    explorer_id="desglosado_explorer",
                    case=False, 
                    multiselect_columns=multiselect_filter_cols,
                    fecha_columns=fecha_columns,
                    numeric_columns=numeric_columns,
                    text_columns=text_columns,
                    excluded_filter_columns=excluded_columns,  # Excluir columnas especificadas
                    container=st.sidebar  # Mostrar los filtros en la barra lateral
                )
                # Update the session state with filtered desglosado data
                st.session_state.df_desglosado = filtered_df_renamed.copy()
        except Exception as e:
            st.sidebar.error(f"Error al aplicar filtros: {e}")
            # Fallback: filtered_df_renamed remains display_data_renamed.copy() (all data from initial load)
            # This was set before the tabs were created.

        # ---- Crear diccionario de configuración de columnas ----
        column_config_dict = {}

        # Configuración para columnas de moneda
        for col in currency_columns:
            if col in filtered_df_renamed.columns: # Verificar que la columna exista en el DF filtrado/renombrado
                column_config_dict[col] = st.column_config.NumberColumn(
                    label=col, # Usar el nombre actual de la columna como etiqueta
                    format="dollar",
                    help=f"Valores monetarios en {col}"
                )

        # Configuración para columnas de URL
        for col in URL_columns:
             if col in filtered_df_renamed.columns: # Verificar que la columna exista
                 column_config_dict[col] = st.column_config.LinkColumn(
                     label=col, # Usar el nombre actual de la columna como etiqueta
                     display_text="Ver PDF", # Texto que se mostrará en el enlace
                     help=f"Enlace al documento PDF ({col})",
                     width="small"
                 )
        # -------------------------------------------------------

        # Mostrar el dataframe filtrado CON configuración de columnas con selección habilitada
        selection_event_desglosado = st.dataframe(
            filtered_df_renamed,
            use_container_width=True,
            column_config=column_config_dict, # Aplicar la configuración
            height=525,  # Aumentar la altura para aprovechar el espacio
            on_select="rerun",
            selection_mode="multi-row"
        )

        # Información sobre el número de filas mostradas con estilo mejorado
        st.info(f"📊 Mostrando {len(filtered_df_renamed):,} de {len(display_data_renamed):,} registros según los filtros aplicados")
        
        # Mostrar las filas guardadas en la selección temporal correspondientes a Desglosado
        if not st.session_state.saved_selections_desglosado.empty:
            st.subheader("Registro Temporal")
            # Agregar mensaje con el número total de filas guardadas
            num_filas = len(st.session_state.saved_selections_desglosado)
            st.dataframe(
                st.session_state.saved_selections_desglosado,
                use_container_width=True,
                column_config=column_config_dict,
                height=300
            )
            st.info(f"📊 Mostrando {num_filas:,} conceptos seleccionados en Desglosado")

    # Pestaña 1 - Vista Concentrada (agrupada por xml_uuid)
    with tab1:
        data_contabilidad_for_tab2 = pd.DataFrame() # Initialize as empty

        # Ensure tab1's filtered data (filtered_df_renamed) is available and has the 'UUID' column
        # Also ensure base data_contabilidad (from main search) is available and has 'xml_uuid'
        if not filtered_df_renamed.empty and "UUID" in filtered_df_renamed.columns and \
           not data_contabilidad.empty and "UUID" in data_contabilidad.columns:
            
            unique_uuids_from_tab1 = filtered_df_renamed["UUID"].unique()

            # Filter data_contabilidad based on the UUIDs present in tab1's filtered_df_renamed
            data_contabilidad_for_tab2 = data_contabilidad[data_contabilidad['UUID'].isin(unique_uuids_from_tab1)].copy()

        # Now, use data_contabilidad_for_tab2 to prepare filtered_concentrado for display
        if not data_contabilidad_for_tab2.empty:
            filtered_concentrado = data_contabilidad_for_tab2 # It's already a copy and filtered by tab1's UUIDs
            filtered_concentrado.rename(columns=column_mapping, inplace=True, errors='ignore')
            # Update the session state with this filtered concentrado
            st.session_state.df_concentrado = filtered_concentrado.copy()
        else:
            # Provide context if no data is shown in tab2
            if filtered_df_renamed.empty or "UUID" not in filtered_df_renamed.columns:
                st.info("No hay datos en la vista Desglosado para filtrar la vista Concentrado, o la columna 'UUID' falta en Desglosado.")
            elif data_contabilidad.empty or "UUID" not in data_contabilidad.columns:
                st.info("No hay datos de contabilidad base para filtrar, o la columna 'UUID' falta en Contabilidad.")
            else:
                # This means UUIDs might have been found in tab1, and data_contabilidad exists, but no matches after filtering.
                st.info("No hay datos de contabilidad que coincidan con los UUIDs de la vista Desglosado.")
            filtered_concentrado = pd.DataFrame() # Ensure it's an empty DataFrame
            st.session_state.df_concentrado = filtered_concentrado.copy() # Ensure session state is updated with empty DataFrame
            
        # Crear diccionario de configuración de columnas para la vista concentrada
        concentrado_config_dict = {}
            
        # Configuración para columnas de moneda
        for col in currency_columns:
            if col in filtered_concentrado.columns:
                    concentrado_config_dict[col] = st.column_config.NumberColumn(
                        label=col,
                        format="dollar",
                        help=f"Valores monetarios en {col} (agrupados)"
                    )
            
        # Configuración para columnas de URL
        for col in URL_columns:
            if col in filtered_concentrado.columns:
                concentrado_config_dict[col] = st.column_config.LinkColumn(
                    label=col,
                    display_text="Ver PDF",
                        help=f"Enlace al documento PDF ({col})",
                        width="small"
                    )

        # Inicializar variables en session state si no existen
        if 'saved_selections' not in st.session_state:
            st.session_state.saved_selections = pd.DataFrame()
        # Inicializar variable para guardar selecciones del dataframe desglosado
        if 'saved_selections_desglosado' not in st.session_state:
            st.session_state.saved_selections_desglosado = pd.DataFrame()
        
        # Mostrar el dataframe concentrado con selección de filas habilitada
        selection_event = st.dataframe(
            filtered_concentrado,
            use_container_width=True,
            column_config=concentrado_config_dict,
            height=525,
            on_select="rerun",
            selection_mode="multi-row"
        )

        st.info(f"📊 Mostrando {len(filtered_concentrado):,} de {len(data_contabilidad):,} facturas")

        # Mostrar las filas guardadas en la selección temporal correspondientes a Concentrado
        if not st.session_state.saved_selections.empty:
            st.subheader("Registro Temporal")
            # Agregar mensaje con el número total de filas guardadas
            num_filas = len(st.session_state.saved_selections)
            st.dataframe(
                st.session_state.saved_selections,
                use_container_width=True,
                column_config=concentrado_config_dict,
                height=300
            )
            st.info(f"📊 Mostrando {num_filas:,} facturas seleccionadas en Concentrado")


    # Mover los botones de selección a la barra lateral
    # Sección de selección de filas en la barra lateral
    st.sidebar.subheader("Herramientas de selección")
    
    # Botones de selección en la misma fila
    col_save, col_clear = st.sidebar.columns(2)
    
    # Botón para guardar selección
    if col_save.button("💾 Guardar filas seleccionadas", key="save_selection"):
        concentrado_seleccionado = False
        desglosado_seleccionado = False
        
        # Verificar si hay selecciones en el dataframe concentrado
        if selection_event.selection and len(selection_event.selection['rows']) > 0:
            # Obtener los índices de las filas seleccionadas
            selected_indices = selection_event.selection['rows']
            # Filtrar el dataframe para obtener solo las filas seleccionadas
            selected_rows = filtered_concentrado.iloc[selected_indices]
            
            # Guardar las filas seleccionadas en session state
            # Si ya hay filas guardadas, concatenar con las nuevas
            if not st.session_state.saved_selections.empty:
                # Concatenar sin duplicados (basado en índices o alguna columna única)
                combined = pd.concat([st.session_state.saved_selections, selected_rows])
                # Eliminar duplicados si existe alguna columna que sea identificador único
                if 'xml_uuid' in combined.columns:
                    st.session_state.saved_selections = combined.drop_duplicates(subset=['xml_uuid'])
                else:
                    st.session_state.saved_selections = combined.drop_duplicates()
            else:
                st.session_state.saved_selections = selected_rows
                
            concentrado_seleccionado = True
        
        # Verificar si hay selecciones en el dataframe desglosado
        if selection_event_desglosado.selection and len(selection_event_desglosado.selection['rows']) > 0:
            # Obtener los índices de las filas seleccionadas
            selected_indices_desglosado = selection_event_desglosado.selection['rows']
            # Filtrar el dataframe para obtener solo las filas seleccionadas
            selected_rows_desglosado = filtered_df_renamed.iloc[selected_indices_desglosado]
            
            # Guardar las filas seleccionadas en session state
            # Si ya hay filas guardadas, concatenar con las nuevas
            if not st.session_state.saved_selections_desglosado.empty:
                # Concatenar sin duplicados (basado en índices o alguna columna única)
                combined = pd.concat([st.session_state.saved_selections_desglosado, selected_rows_desglosado])
                # Eliminar duplicados si existe alguna columna que sea identificador único
                if 'xml_uuid' in combined.columns:
                    st.session_state.saved_selections_desglosado = combined.drop_duplicates(subset=['xml_uuid'])
                else:
                    st.session_state.saved_selections_desglosado = combined.drop_duplicates()
            else:
                st.session_state.saved_selections_desglosado = selected_rows_desglosado
                
            desglosado_seleccionado = True
        
        # Mostrar mensaje de éxito o advertencia según corresponda
        if concentrado_seleccionado or desglosado_seleccionado:
            mensaje = []
            if concentrado_seleccionado:
                mensaje.append(f"{len(selection_event.selection['rows'])} facturas de Concentrado")
            if desglosado_seleccionado:
                mensaje.append(f"{len(selection_event_desglosado.selection['rows'])} conceptos de Desglosado")
                
            st.session_state.sidebar_toast_message = f"Guardadas: {' y '.join(mensaje)}"
            st.session_state.sidebar_toast_icon = '✅'  # Icono de éxito (check verde)
            st.rerun()
        else:
            st.session_state.sidebar_toast_message = "No hay filas seleccionadas para guardar"
            st.session_state.sidebar_toast_icon = '⚠️'  # Icono de advertencia
            st.rerun()

    # Botón para limpiar selección
    if col_clear.button("🗑️ Limpiar filas seleccionadas", key="clear_selection"):
        st.session_state.saved_selections = pd.DataFrame()
        st.session_state.saved_selections_desglosado = pd.DataFrame()
        st.session_state.sidebar_toast_message = "Tablas temporales limpiadas correctamente"
        st.session_state.sidebar_toast_icon = '✅'  # Icono de éxito (check verde)
        st.rerun()
        
    # Separador para filtros automáticos
    st.sidebar.markdown("---")
    st.sidebar.subheader("Filtros automáticos")
    
    # Botón para filtrar y guardar facturas con descuento > 0
    col1, col2 = st.sidebar.columns(2)
    
    if col2.button("🔍 Facturas con descuento", key="filter_descuento"):
        try:
            # Variables para contar facturas encontradas
            num_facturas_concentrado = 0
            num_facturas_desglosado = 0
            
            # Filtrar en dataframe concentrado
            if 'Descuento' in filtered_concentrado.columns:
                # Filtrar facturas con descuento > 0
                df_con_descuento = filtered_concentrado[filtered_concentrado['Descuento'] > 0].copy()
                num_facturas_concentrado = len(df_con_descuento)
                
                if num_facturas_concentrado > 0:
                    # Guardar en el dataframe temporal
                    if not st.session_state.saved_selections.empty:
                        # Concatenar con selecciones existentes
                        combined = pd.concat([st.session_state.saved_selections, df_con_descuento])
                        # Eliminar duplicados si existe una columna identificadora única
                        if 'xml_uuid' in combined.columns:
                            st.session_state.saved_selections = combined.drop_duplicates(subset=['xml_uuid'])
                        else:
                            st.session_state.saved_selections = combined.drop_duplicates()
                    else:
                        # Guardar directamente si no hay selecciones previas
                        st.session_state.saved_selections = df_con_descuento
            
            # Filtrar en dataframe desglosado
            if 'Descuento' in filtered_df_renamed.columns:
                # Filtrar facturas con descuento > 0
                df_desglosado_con_descuento = filtered_df_renamed[filtered_df_renamed['Descuento'] > 0].copy()
                num_facturas_desglosado = len(df_desglosado_con_descuento)
                
                if num_facturas_desglosado > 0:
                    # Guardar en el dataframe temporal
                    if not st.session_state.saved_selections_desglosado.empty:
                        # Concatenar con selecciones existentes
                        combined = pd.concat([st.session_state.saved_selections_desglosado, df_desglosado_con_descuento])
                        # Eliminar duplicados si existe una columna identificadora única
                        if 'xml_uuid' in combined.columns:
                            st.session_state.saved_selections_desglosado = combined.drop_duplicates(subset=['xml_uuid'])
                        else:
                            st.session_state.saved_selections_desglosado = combined.drop_duplicates()
                    else:
                        # Guardar directamente si no hay selecciones previas
                        st.session_state.saved_selections_desglosado = df_desglosado_con_descuento
            
            # Mostrar mensaje según resultados
            if num_facturas_concentrado > 0 or num_facturas_desglosado > 0:
                mensaje = []
                if num_facturas_concentrado > 0:
                    mensaje.append(f":blue[**{num_facturas_concentrado} facturas**]")
                if num_facturas_desglosado > 0:
                    mensaje.append(f":green[**{num_facturas_desglosado} conceptos**]")
                st.session_state.sidebar_toast_message = f"Se guardaron: {' y '.join(mensaje)} *con descuento*"
                st.session_state.sidebar_toast_icon = '✅'  # Icono de éxito
                st.rerun()
            else:
                st.session_state.sidebar_toast_message = "No se encontraron facturas con descuento en ninguna vista"
                st.session_state.sidebar_toast_icon = '⚠️'  # Icono de advertencia
                st.rerun()
        except Exception as e:
            st.session_state.sidebar_toast_message = f"Error al filtrar: {e}"
            st.session_state.sidebar_toast_icon = '🚨'  # Icono de error
            st.rerun()
    
    # Botón para filtrar y guardar facturas con retenciones (ISH, Retención IVA, Retención ISR)
    if col1.button("🔍 Facturas con retenciones", key="filter_retenciones"):
        try:
            # Variables para contar facturas encontradas
            num_facturas_concentrado = 0
            num_facturas_desglosado = 0
            
            # Filtrar en dataframe concentrado
            columnas_retenciones = ['ISH', 'Retención IVA', 'Retención ISR']
            
            # Concentrado
            columnas_existentes_concentrado = [col for col in columnas_retenciones if col in filtered_concentrado.columns]
            if len(columnas_existentes_concentrado) > 0:
                # Crear máscara para cada columna existente
                mask = pd.Series([False] * len(filtered_concentrado), index=filtered_concentrado.index)
                
                # Sumar las máscaras para cada columna
                for col in columnas_existentes_concentrado:
                    mask = mask | (filtered_concentrado[col] != 0)
                
                # Filtrar las facturas que cumplen con al menos una condición
                df_con_retenciones = filtered_concentrado[mask].copy()
                num_facturas_concentrado = len(df_con_retenciones)
                
                if num_facturas_concentrado > 0:
                    # Guardar en el dataframe temporal
                    if not st.session_state.saved_selections.empty:
                        # Concatenar con selecciones existentes
                        combined = pd.concat([st.session_state.saved_selections, df_con_retenciones])
                        # Eliminar duplicados si existe una columna identificadora única
                        if 'xml_uuid' in combined.columns:
                            st.session_state.saved_selections = combined.drop_duplicates(subset=['xml_uuid'])
                        else:
                            st.session_state.saved_selections = combined.drop_duplicates()
                    else:
                        # Guardar directamente si no hay selecciones previas
                        st.session_state.saved_selections = df_con_retenciones
            
            # Desglosado
            columnas_existentes_desglosado = [col for col in columnas_retenciones if col in filtered_df_renamed.columns]
            if len(columnas_existentes_desglosado) > 0:
                # Crear máscara para cada columna existente
                mask = pd.Series([False] * len(filtered_df_renamed), index=filtered_df_renamed.index)
                
                # Sumar las máscaras para cada columna
                for col in columnas_existentes_desglosado:
                    mask = mask | (filtered_df_renamed[col] != 0)
                
                # Filtrar las facturas que cumplen con al menos una condición
                df_desglosado_con_retenciones = filtered_df_renamed[mask].copy()
                num_facturas_desglosado = len(df_desglosado_con_retenciones)
                
                if num_facturas_desglosado > 0:
                    # Guardar en el dataframe temporal
                    if not st.session_state.saved_selections_desglosado.empty:
                        # Concatenar con selecciones existentes
                        combined = pd.concat([st.session_state.saved_selections_desglosado, df_desglosado_con_retenciones])
                        # Eliminar duplicados si existe una columna identificadora única
                        if 'xml_uuid' in combined.columns:
                            st.session_state.saved_selections_desglosado = combined.drop_duplicates(subset=['xml_uuid'])
                        else:
                            st.session_state.saved_selections_desglosado = combined.drop_duplicates()
                    else:
                        # Guardar directamente si no hay selecciones previas
                        st.session_state.saved_selections_desglosado = df_desglosado_con_retenciones
            
            # Mostrar mensaje según resultados
            if num_facturas_concentrado > 0 or num_facturas_desglosado > 0:
                mensaje = []
                if num_facturas_concentrado > 0:
                    mensaje.append(f":blue[**{num_facturas_concentrado} facturas**]")
                if num_facturas_desglosado > 0:
                    mensaje.append(f":green[**{num_facturas_desglosado} conceptos**]")
                st.session_state.sidebar_toast_message = f"Se guardaron: {' y '.join(mensaje)} *con retenciones*"
                st.session_state.sidebar_toast_icon = '✅'  # Icono de éxito
                st.rerun()
            else:
                st.session_state.sidebar_toast_message = "No se encontraron facturas con retenciones en ninguna vista"
                st.session_state.sidebar_toast_icon = '⚠️'  # Icono de advertencia
                st.rerun()
        except Exception as e:
            st.session_state.sidebar_toast_message = f"Error al filtrar: {e}"
            st.session_state.sidebar_toast_icon = '🚨'  # Icono de error
            st.rerun()
    
    # Botón para filtrar y guardar facturas con Tasa 0 y facturas en USD
    col3, col4 = st.sidebar.columns(2)
    if col3.button("🔍 Facturas con Impuesto Tasa 0", key="filter_tasa0"):
        try:
            # Variables para contar facturas encontradas
            num_facturas_concentrado = 0
            num_facturas_desglosado = 0
            columnas_tasas = ['Venta Tasa 0%', 'Venta Tasa 16%']
            
            # Filtrar en Concentrado
            columnas_existentes_concentrado = [col for col in columnas_tasas if col in filtered_concentrado.columns]
            if len(columnas_existentes_concentrado) > 0:
                # Crear máscaras para las condiciones
                mask_total = pd.Series([False] * len(filtered_concentrado), index=filtered_concentrado.index)
                
                # Condición 1: Venta Tasa 0% > 0
                if 'Venta Tasa 0%' in filtered_concentrado.columns:
                    mask_tasa0_mayor_0 = filtered_concentrado['Venta Tasa 0%'] > 0
                    mask_total = mask_total | mask_tasa0_mayor_0
                
                # Condición 2: Ambas tasas son 0
                if all(col in filtered_concentrado.columns for col in columnas_tasas):
                    mask_ambas_0 = (filtered_concentrado['Venta Tasa 0%'] == 0) & (filtered_concentrado['Venta Tasa 16%'] == 0)
                    mask_total = mask_total | mask_ambas_0
                
                # Filtrar las facturas que cumplen con alguna condición
                df_tasa0 = filtered_concentrado[mask_total].copy()
                num_facturas_concentrado = len(df_tasa0)
                
                if num_facturas_concentrado > 0:
                    # Guardar en el dataframe temporal
                    if not st.session_state.saved_selections.empty:
                        # Concatenar con selecciones existentes
                        combined = pd.concat([st.session_state.saved_selections, df_tasa0])
                        # Eliminar duplicados
                        if 'xml_uuid' in combined.columns:
                            st.session_state.saved_selections = combined.drop_duplicates(subset=['xml_uuid'])
                        else:
                            st.session_state.saved_selections = combined.drop_duplicates()
                    else:
                        # Guardar directamente si no hay selecciones previas
                        st.session_state.saved_selections = df_tasa0
            
            # Filtrar en Desglosado
            columnas_existentes_desglosado = [col for col in columnas_tasas if col in filtered_df_renamed.columns]
            if len(columnas_existentes_desglosado) > 0:
                # Crear máscaras para las condiciones
                mask_total = pd.Series([False] * len(filtered_df_renamed), index=filtered_df_renamed.index)
                
                # Condición 1: Venta Tasa 0% > 0
                if 'Venta Tasa 0%' in filtered_df_renamed.columns:
                    mask_tasa0_mayor_0 = filtered_df_renamed['Venta Tasa 0%'] > 0
                    mask_total = mask_total | mask_tasa0_mayor_0
                
                # Condición 2: Ambas tasas son 0
                if all(col in filtered_df_renamed.columns for col in columnas_tasas):
                    mask_ambas_0 = (filtered_df_renamed['Venta Tasa 0%'] == 0) & (filtered_df_renamed['Venta Tasa 16%'] == 0)
                    mask_total = mask_total | mask_ambas_0
                
                # Filtrar las facturas que cumplen con alguna condición
                df_desglosado_tasa0 = filtered_df_renamed[mask_total].copy()
                num_facturas_desglosado = len(df_desglosado_tasa0)
                
                if num_facturas_desglosado > 0:
                    # Guardar en el dataframe temporal
                    if not st.session_state.saved_selections_desglosado.empty:
                        # Concatenar con selecciones existentes
                        combined = pd.concat([st.session_state.saved_selections_desglosado, df_desglosado_tasa0])
                        # Eliminar duplicados
                        if 'xml_uuid' in combined.columns:
                            st.session_state.saved_selections_desglosado = combined.drop_duplicates(subset=['xml_uuid'])
                        else:
                            st.session_state.saved_selections_desglosado = combined.drop_duplicates()
                    else:
                        # Guardar directamente si no hay selecciones previas
                        st.session_state.saved_selections_desglosado = df_desglosado_tasa0
            
            # Mostrar mensaje según resultados
            if num_facturas_concentrado > 0 or num_facturas_desglosado > 0:
                mensaje = []
                if num_facturas_concentrado > 0:
                    mensaje.append(f":blue[**{num_facturas_concentrado} facturas**]")
                if num_facturas_desglosado > 0:
                    mensaje.append(f":green[**{num_facturas_desglosado} conceptos**]")
                st.session_state.sidebar_toast_message = f"Se guardaron: {' y '.join(mensaje)} *con Tasa 0*"
                st.session_state.sidebar_toast_icon = '✅'  # Icono de éxito
                st.rerun()
            else:
                st.session_state.sidebar_toast_message = "No se encontraron facturas que cumplan las condiciones de Impuesto Tasa 0"
                st.session_state.sidebar_toast_icon = '⚠️'  # Icono de advertencia
                st.rerun()
        except Exception as e:
            st.session_state.sidebar_toast_message = f"Error al filtrar: {e}"
            st.session_state.sidebar_toast_icon = '🚨'  # Icono de error
            st.rerun()
    
    # Botón para filtrar y guardar facturas en moneda USD
    if col4.button("💲 Facturas con moneda en USD", key="filter_usd"):
        try:
            # Variables para contar facturas encontradas
            num_facturas_concentrado = 0
            num_facturas_desglosado = 0
            
            # Filtrar en Concentrado
            if 'Moneda' in filtered_concentrado.columns:
                # Filtrar facturas en USD
                df_usd = filtered_concentrado[filtered_concentrado['Moneda'] == 'USD'].copy()
                num_facturas_concentrado = len(df_usd)
                
                if num_facturas_concentrado > 0:
                    # Guardar en el dataframe temporal
                    if not st.session_state.saved_selections.empty:
                        # Concatenar con selecciones existentes
                        combined = pd.concat([st.session_state.saved_selections, df_usd])
                        # Eliminar duplicados
                        if 'xml_uuid' in combined.columns:
                            st.session_state.saved_selections = combined.drop_duplicates(subset=['xml_uuid'])
                        else:
                            st.session_state.saved_selections = combined.drop_duplicates()
                    else:
                        # Guardar directamente si no hay selecciones previas
                        st.session_state.saved_selections = df_usd
            
            # Filtrar en Desglosado
            if 'Moneda' in filtered_df_renamed.columns:
                # Filtrar facturas en USD
                df_desglosado_usd = filtered_df_renamed[filtered_df_renamed['Moneda'] == 'USD'].copy()
                num_facturas_desglosado = len(df_desglosado_usd)
                
                if num_facturas_desglosado > 0:
                    # Guardar en el dataframe temporal
                    if not st.session_state.saved_selections_desglosado.empty:
                        # Concatenar con selecciones existentes
                        combined = pd.concat([st.session_state.saved_selections_desglosado, df_desglosado_usd])
                        # Eliminar duplicados
                        if 'xml_uuid' in combined.columns:
                            st.session_state.saved_selections_desglosado = combined.drop_duplicates(subset=['xml_uuid'])
                        else:
                            st.session_state.saved_selections_desglosado = combined.drop_duplicates()
                    else:
                        # Guardar directamente si no hay selecciones previas
                        st.session_state.saved_selections_desglosado = df_desglosado_usd
            
            # Mostrar mensaje según resultados
            if num_facturas_concentrado > 0 or num_facturas_desglosado > 0:
                mensaje = []
                if num_facturas_concentrado > 0:
                    mensaje.append(f":blue[**{num_facturas_concentrado} facturas**]")
                if num_facturas_desglosado > 0:
                    mensaje.append(f":green[**{num_facturas_desglosado} conceptos**]")
                st.session_state.sidebar_toast_message = f"Se guardaron: {' y '.join(mensaje)} *en USD*"
                st.session_state.sidebar_toast_icon = '✅'  # Icono de éxito
                st.rerun()
            else:
                st.session_state.sidebar_toast_message = "No se encontraron facturas en USD en ninguna vista"
                st.session_state.sidebar_toast_icon = '⚠️'  # Icono de advertencia
                st.rerun()
        except Exception as e:
            st.session_state.sidebar_toast_message = f"Error al filtrar: {e}"
            st.session_state.sidebar_toast_icon = '🚨'  # Icono de error
            st.rerun()
    
    # Separador para sección de descargas
    st.sidebar.markdown("---")
    st.sidebar.subheader("Descargas")
    
    # Inicializar variables en session_state para el dialogo de descargas
    if 'download_source' not in st.session_state:
        st.session_state.download_source = 'filtered_concentrado'
    if 'download_columns' not in st.session_state:
        st.session_state.download_columns = []
    if 'download_mode' not in st.session_state:
        st.session_state.download_mode = 'combined'
    if 'download_dir' not in st.session_state:
        # Definir un directorio de descarga predeterminado para evitar errores
        fecha_actual = time.strftime('%Y-%m-%d')
        # Verificar que base_downloads_dir exista en session_state
        if 'base_downloads_dir' not in st.session_state:
            # Intentar determinar la carpeta de Descargas del usuario de manera más portátil
            home_dir = os.path.expanduser('~')
            # Comprobar rutas comunes de Descargas/Downloads en diferentes sistemas
            download_candidates = [
                os.path.join(home_dir, 'Downloads'),  # Inglés/Mac/Linux
                os.path.join(home_dir, 'Descargas'),  # Español
                os.path.join(home_dir, 'Desktop', 'Downloads'),  # Alternativa en escritorio
                os.path.join(home_dir, 'Documents')  # Alternativa en documentos
            ]
            
            # Usar la primera ruta que exista o crear la carpeta de Downloads en el directorio home
            for path in download_candidates:
                if os.path.exists(path) and os.path.isdir(path):
                    st.session_state.base_downloads_dir = path
                    break
            else:
                # Si ninguna ruta existe, usar el directorio home
                st.session_state.base_downloads_dir = home_dir
        
        st.session_state.download_dir = os.path.join(st.session_state.base_downloads_dir, f'Facturas_StreamlPT_{fecha_actual}')
    
    # Preparar ruta base de la carpeta de descargas (pero no crearla aún)
    if 'base_downloads_dir' not in st.session_state:
        home_dir = os.path.expanduser('~')
        st.session_state.base_downloads_dir = os.path.join(home_dir, 'Downloads')
        
    if 'download_progress' not in st.session_state:
        st.session_state.download_progress = 0
    if 'download_results' not in st.session_state:
        st.session_state.download_results = None
    
    # Definir la función de diálogo para descargas
    @st.dialog("💾 Descarga de archivos PDF", width="large")
    def configurar_descarga():
        # st.subheader("Configuración de descarga de PDFs")
        
        # Selección de fuente de datos
        st.session_state.download_source = st.radio(
            "Seleccione la fuente de datos:",
            options=[
                'df_concentrado', 
                'saved_selections',
                'df_desglosado',
                'saved_selections_desglosado'
            ],
            format_func=lambda x: {
                'df_concentrado': "Tabla Concentrado", 
                'saved_selections': "Tabla Registro temporal (Concentrado)",
                'saved_selections_desglosado': "Tabla Registro temporal (Desglosado)",
                'df_desglosado': "Tabla Desglosado"
            }[x],
            index=0
        )
        
        # Mapeo de opciones a dataframes
        data_source_map = {
            'df_concentrado': st.session_state.df_concentrado,
            'saved_selections': st.session_state.saved_selections,
            'saved_selections_desglosado': st.session_state.saved_selections_desglosado,
            'df_desglosado': st.session_state.df_desglosado
        }
        
        # Crear una copia del DataFrame para trabajar con él y aplicar ordenamiento
        selected_df = data_source_map[st.session_state.download_source].copy()
        rows_count = len(selected_df) if not selected_df.empty else 0
        
        st.info(f"📊 Fuente seleccionada: {rows_count:,} registros disponibles")
        
        # No mostrar opciones si el dataframe está vacío
        if not selected_df.empty:
            # Selección de columnas para descargar
            available_url_columns = [col for col in URL_columns if col in selected_df.columns]
            
            if available_url_columns:
                st.session_state.download_columns = st.multiselect(
                    "Seleccione las columnas para descargar:",
                    options=available_url_columns,
                    placeholder="Seleccione columnas para descargar",
                    default=available_url_columns[:1] if available_url_columns else []
                )
                
                if st.session_state.download_columns:
                    # Opciones de descarga de PDFs
                    st.session_state.download_mode = st.radio(
                        "Modo de descarga:",
                        options=['combined', 'joined'],
                        format_func=lambda x: {
                            'combined': "PDF por fila", 
                            'joined': "Un solo archivo"
                        }[x],
                        index=0,
                        help="'PDF por fila' combina los documentos de cada fila en un solo PDF. 'Un solo archivo' une todos los PDFs en un único archivo."
                    )
                    
                    # # Mostrar información sobre el destino y método de descarga
                    # st.subheader("💾 Descarga de archivos", divider="rainbow")
                    
                    # # Mantener la variable en session_state por compatibilidad
                    # if 'download_dir' not in st.session_state:
                    #     st.session_state.download_dir = "browser_download"
                    
                    
                    # # Una sola línea informativa según el modo seleccionado
                    # if st.session_state.download_mode == 'combined':
                    #     st.info("🗞️ Se creará un PDF por cada fila")
                    # elif st.session_state.download_mode == 'joined':
                    #     st.info("🗞️ Se creará un único archivo PDF con todos los documentos")
                    
                    # Añadir sistema de ordenamiento personalizado
                    st.subheader("Orden de descarga", divider="rainbow")
                    
                    # Obtener todas las columnas del DataFrame para el multiselect
                    all_columns = selected_df.columns.tolist()
                    
                    # Selección de columnas para ordenar
                    sort_columns = st.multiselect(
                        "Seleccione columnas para ordenar (en orden de prioridad):",
                        options=all_columns,
                        placeholder="Elige el orden para la descarga (columnas)",
                        help="Las columnas se ordenarán en el orden en que las seleccione"
                    )
                    
                    # Si hay columnas seleccionadas, mostrar controles de dirección de ordenamiento
                    sort_directions = {}
                    if sort_columns:
                        # st.write("Seleccione la dirección de ordenamiento para cada columna:")
                        cols = st.columns(min(len(sort_columns), 3))  # Máximo 3 columnas por fila
                        
                        for i, column in enumerate(sort_columns):
                            col_idx = i % 3  # Distribuir en las columnas disponibles
                            with cols[col_idx]:
                                direction = st.segmented_control(
                                    label=f"{column}",
                                    options=["Ascendente", "Descendente"],
                                    key=f"sort_dir_{i}"
                                )
                                sort_directions[column] = True if direction == "Ascendente" else False
                    
                    # Botón para iniciar la descarga
                    if st.button("Iniciar descarga", key="start_download", type="primary"):
                        # Ya no necesitamos crear directorios locales, los archivos se descargaru00e1n directamente
                        fecha_actual = time.strftime('%Y-%m-%d')
                        
                        # Usaremos memoria para almacenar los PDFs generados
                        st.session_state.generated_pdfs = []
                        st.session_state.combined_pdf_bytes = None
                        
                        # Aplicar el ordenamiento al dataframe según las selecciones del usuario
                        df_to_process = selected_df.copy()
                        
                        # Verificar si estamos utilizando el dataframe desglosado y filtrar URLs únicas si es necesario
                        if st.session_state.download_source == 'saved_selections_desglosado' or st.session_state.download_source == 'df_desglosado' and st.session_state.download_columns:
                            st.info("Filtrando URLs únicas para evitar duplicados...")
                            
                            # Verificar qué columnas de URL están seleccionadas para descarga
                            url_cols_to_process = [col for col in st.session_state.download_columns if col in URL_columns]
                            
                            if url_cols_to_process and not df_to_process.empty:
                                # Crear un nuevo DataFrame para guardar filas con URLs únicas
                                unique_df = pd.DataFrame(columns=df_to_process.columns)
                                unique_urls = set()
                                
                                for _, row in df_to_process.iterrows():
                                    row_urls = []
                                    # Recopilar URLs de esta fila
                                    for col in url_cols_to_process:
                                        if pd.notna(row.get(col)) and row.get(col) and isinstance(row.get(col), str):
                                            row_urls.append(row.get(col))
                                    
                                    # Verificar si ya tenemos estas URLs
                                    new_url_found = False
                                    for url in row_urls:
                                        if url not in unique_urls:
                                            unique_urls.add(url)
                                            new_url_found = True
                                    
                                    # Si hay una URL nueva, guardar esta fila
                                    if new_url_found:
                                        unique_df = pd.concat([unique_df, pd.DataFrame([row])], ignore_index=True)
                                
                                # Reemplazar el DataFrame original con el filtrado
                                filtered_count = len(df_to_process) - len(unique_df)
                                df_to_process = unique_df
                                if filtered_count > 0:
                                    st.info(f"Se eliminaron {filtered_count} filas con URLs duplicadas")
                        
                        if sort_columns:
                            # Aplicar sort con las columnas y direcciones seleccionadas
                            st.info(f"🔄 Ordenando por: {', '.join(sort_columns)}")
                            df_to_process = df_to_process.sort_values(
                                by=sort_columns,
                                ascending=[sort_directions[col] for col in sort_columns]
                            ).reset_index(drop=True)
                        
                        # Inicializar lista para almacenar datos de PDFs
                        if 'generated_pdfs' not in st.session_state:
                            st.session_state.generated_pdfs = []
                        
                        # Limpiar lista anterior
                        st.session_state.generated_pdfs = []
                        
                        # Inicializar variable para control de cancelación
                        if 'download_cancelled' not in st.session_state:
                            st.session_state.download_cancelled = False
                        st.session_state.download_cancelled = False
                        
                        gestor = GestorDescargas(max_workers=10)
                        
                        # Añadir descargas a la cola
                        file_count = 0
                        
                        # Procesar según modo seleccionado - solo 'combined' y 'joined'
                        # Código común para ambos modos
                        # Crear barra de progreso para mostrar el avance del proceso
                        progress_bar = st.progress(0)
                        progress_text = st.empty()
                        
                        # Crear un botón de cancelación
                        cancel_col = st.empty()
                        if cancel_col.button("❌ Cancelar descarga", key="cancel_download", type="secondary"):
                            st.session_state.download_cancelled = True
                            progress_text.write("Cancelando descarga...")
                            time.sleep(1)
                            st.rerun()
                        
                        if st.session_state.download_mode == 'combined':
                            progress_text.write("Procesando documentos para combinarlos por fila...")
                        else:  # modo 'joined'
                            progress_text.write("Procesando documentos para combinarlos en un solo archivo...")
                            
                        # Verificar disponibilidad de columnas 'Obra' y 'Proveedor' para nombrado
                        nombre_cols_disponibles = all(col in df_to_process.columns for col in ['Obra', 'Proveedor'])
                        if not nombre_cols_disponibles:
                            st.warning("⚠️ No se encontraron las columnas 'Obra' y 'Proveedor'. Se usarán nombres genéricos.")

                        # Crear gestor para combinar PDFs
                        combinador = CombinadorPDF()
                        if not combinador.can_combine:
                            st.error("PyPDF2 no está instalado. Por favor, ejecuta 'pip install PyPDF2' y reinicia la aplicación.")
                            return
                                
                        # Contador para filas procesadas y nombre de archivos
                        counter = 0
                        total_docs = 0
                        combined_results = []
                        # Ya no necesitamos all_combined_pdfs ya que guardamos todo en memoria
                        
                        # Obtener el total de filas para calcular el progreso
                        total_rows = len(df_to_process)
                        if total_rows == 0:
                            progress_bar.progress(100)  # No hay filas, progreso completo
                            progress_text.write("No hay documentos para procesar")
                        else:
                            progress_text.write(f"Procesando {total_rows} filas de documentos...")
                            
                        # Procesar cada fila y combinar sus documentos en un solo PDF
                        for idx, row in df_to_process.iterrows():
                            # Verificar si se ha cancelado la descarga
                            if st.session_state.download_cancelled:
                                progress_text.write("Descarga cancelada por el usuario")
                                break
                                
                            counter += 1
                            
                            # Actualizar la barra de progreso - fase 1 (50% del proceso total)
                            # La primera mitad del progreso es para el procesamiento de filas
                            progress_percent = min(counter / total_rows * 50, 50)
                            progress_bar.progress(int(progress_percent))
                            progress_text.write(f"Procesando fila {counter} de {total_rows}...")
                            
                            # Recopilar URLs a combinar para esta fila (solo URLs únicas)
                            urls = []
                            urls_set = set()  # Para evitar duplicados dentro de una misma fila
                            for col in st.session_state.download_columns:
                                if pd.notna(row.get(col)) and row.get(col) and isinstance(row.get(col), str):
                                    url = row.get(col)
                                    if url not in urls_set:  # Solo agregar si no está ya en la lista
                                        urls_set.add(url)
                                        urls.append(url)
                                
                            if not urls:
                                continue  # No hay URLs en esta fila, continuar con la siguiente
                            
                            # Generar nombre de archivo basado en Obra y Proveedor (si están disponibles)
                            if nombre_cols_disponibles:
                                obra = sanitizar_nombre_archivo(str(row.get('Obra', '')))
                                proveedor = sanitizar_nombre_archivo(str(row.get('Proveedor', '')))
                                filename = f"{counter}_{obra}_{proveedor}.pdf"
                            else:
                                # Usar UUID o fecha + contador como identificador único
                                if 'xml_uuid' in row:
                                    filename = f"{counter}_{sanitizar_nombre_archivo(str(row.get('xml_uuid')))}.pdf"
                                else:
                                    filename = f"{counter}_combinado_{time.strftime('%H%M%S')}_{idx}.pdf"
                            
                            # Combinar los PDFs y guardarlos en memoria
                            try:
                                # Descargar todos los PDFs a memoria
                                pdf_data = io.BytesIO()
                                exito, mensaje, pdf_bytes = combinador.combinar_pdfs_a_memoria(urls)
                                
                                if exito and pdf_bytes:
                                    # Guardar los datos del PDF para su descarga posterior
                                    st.session_state.generated_pdfs.append({
                                        'nombre': filename,
                                        'datos': pdf_bytes,
                                        'urls': urls,
                                        'index': counter
                                    })
                                    exito = True
                                    mensaje = f"PDF {filename} preparado para descarga"
                                else:
                                    exito = False
                                    mensaje = f"Error al combinar PDFs: {mensaje}"
                            except Exception as e:
                                exito = False
                                mensaje = f"Error al procesar PDFs: {str(e)}"
                            combined_results.append({
                                'exito': exito,
                                'mensaje': mensaje,
                                'urls': urls
                            })
                            
                            if exito:
                                total_docs += 1
                                # No necesitamos hacer nada más para el modo 'joined'
                                # porque los PDFs ya están guardados en st.session_state.generated_pdfs
                        
                        # Actualizar progreso al 50% después de procesar todas las filas
                        progress_bar.progress(50)
                        progress_text.write("Fase 1 completada: Procesamiento de documentos")
                        
                        # Mostrar resultados en formato de tabla
                        if combined_results:
                            # Si estamos en modo 'joined', combinar todos los PDFs en uno solo
                            if st.session_state.download_mode == 'joined' and st.session_state.generated_pdfs:
                                progress_text.write("Uniendo todos los PDFs en un solo archivo...")
                                # Crear un PDF combinado en memoria a partir de los PDFs individuales
                                timestamp = time.strftime('%Y%m%d_%H%M%S')
                                nombre_final = f"concentrado_{timestamp}.pdf"
                                
                                progress_text.write("Combinando todos los PDFs en memoria...")
                                
                                # Usar BytesIO para trabajar en memoria
                                pdf_final = io.BytesIO()
                                # Usaremos el mismo combinador de PDFs que ya tenemos instanciado
                                archivos_unidos = 0
                                total_archivos = len(st.session_state.generated_pdfs)
                                
                                try:
                                    # Extraer los bytes de cada PDF para pasarlos al combinador
                                    lista_pdfs_bytes = []
                                    for i, pdf_info in enumerate(st.session_state.generated_pdfs):
                                        # Verificar si se ha cancelado la descarga
                                        if st.session_state.download_cancelled:
                                            progress_text.write("Descarga cancelada por el usuario")
                                            break
                                                
                                        # Usar directamente los bytes almacenados
                                        lista_pdfs_bytes.append(io.BytesIO(pdf_info['datos']))
                                        
                                        # Actualizar progreso - fase 2 (del 50% al 90%)
                                        progress_percent = 50 + int((i + 1) / total_archivos * 40)
                                        progress_bar.progress(progress_percent, text=f"Preparando PDFs ({i+1}/{total_archivos})")
                                    
                                    # Crear una lista temporal de URLs falsas para usar con combinar_pdfs_a_memoria
                                    # Esto es un hack para reutilizar la función existente, que espera URLs
                                    progress_text.write("Combinando PDFs en memoria...")
                                    
                                    # En lugar de descargar de URLs, vamos a abrir directamente los bytes con pikepdf
                                    # y combinarlos manualmente como lo hace la función combinar_pdfs_a_memoria
                                    import pikepdf
                                    
                                    # Crear un nuevo PDF
                                    pdf_final_obj = pikepdf.Pdf.new()
                                    
                                    # Añadir cada PDF a la combinación
                                    for i, pdf_bytes_io in enumerate(lista_pdfs_bytes):
                                        try:
                                            # Abrir el PDF desde los bytes
                                            pdf = pikepdf.Pdf.open(pdf_bytes_io)
                                            # Añadir todas sus páginas al PDF final
                                            pdf_final_obj.pages.extend(pdf.pages)
                                            archivos_unidos += 1
                                        except Exception as e:
                                            progress_text.write(f"⚠️ No se pudo unir el PDF #{i+1}: {str(e)}")
                                    
                                    # Guardar el PDF combinado en memoria
                                    pdf_final.seek(0)
                                    pdf_final_obj.save(pdf_final)
                                    pdf_final.seek(0)
                                    pdf_bytes = pdf_final.read()
                                    
                                    # Cerrar el objeto PDF
                                    pdf_final_obj.close()
                                    
                                    # Guardar los bytes para el botón de descarga
                                    st.session_state.combined_pdf_bytes = pdf_bytes
                                    st.session_state.combined_pdf_name = nombre_final
                                    
                                    # Actualizar progreso - fase final (100%)
                                    progress_bar.progress(100, text=f"¡Proceso completado! {archivos_unidos} PDFs unidos")
                                    progress_text.write(f"✅ PDFs combinados correctamente. Listo para descargar.")
                                    
                                    # Crear un ZIP que contenga el PDF combinado
                                    zip_buffer = io.BytesIO()
                                    timestamp = time.strftime('%Y%m%d_%H%M%S')
                                    zip_filename = f"documento_combinado_{timestamp}.zip"
                                    
                                    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
                                        zip_file.writestr(nombre_final, pdf_bytes)
                                    
                                    # Preparar el buffer para descarga
                                    zip_buffer.seek(0)
                                    zip_data = zip_buffer.getvalue()
                                    
                                    # Botón de descarga del ZIP
                                    st.download_button(
                                        label=f"💾 Descargar PDF combinado en ZIP",
                                        data=zip_data,
                                        file_name=zip_filename,
                                        mime="application/zip",
                                        key="download_combined_zip",
                                        help="Haga clic para descargar el PDF combinado en un archivo ZIP"
                                    )
                                    
                                except Exception as e:
                                    progress_text.write(f"❌ Error al unir PDFs: {str(e)}")
                                    st.error(f"Error al combinar todos los PDFs: {e}")
                                        
                            else:  # modo 'combined'
                                # Completar la barra de progreso para modo 'combined'
                                progress_bar.progress(100)
                                progress_text.write(f"Completado: {total_docs} PDFs generados")
                                
                                # Mensaje de éxito para modo 'combined' (PDF por fila)
                                st.success(f"✅ Proceso finalizado: {total_docs} PDFs generados de {counter} filas procesadas")
                                
                                # Crear un ZIP con todos los PDFs individuales
                                zip_buffer = io.BytesIO()
                                timestamp = time.strftime('%Y%m%d_%H%M%S')
                                zip_filename = f"documentos_pdf_{timestamp}.zip"
                                
                                with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
                                    for pdf_info in st.session_state.generated_pdfs:
                                        zip_file.writestr(pdf_info['nombre'], pdf_info['datos'])
                                
                                # Preparar el buffer para descarga
                                zip_buffer.seek(0)
                                zip_data = zip_buffer.getvalue()
                                
                                # Botón de descarga del ZIP completo
                                st.download_button(
                                    label=f"💾 Descargar todos los PDFs en un ZIP ({len(st.session_state.generated_pdfs)} archivos)",
                                    data=zip_data,
                                    file_name=zip_filename,
                                    mime="application/zip",
                                    key="download_all_zip",
                                    help="Haga clic para descargar todos los PDFs en un solo archivo ZIP"
                                )

                            # Mostrar errores solo si existen
                            errores = [res for res in combined_results if not res['exito']]
                            if errores:
                                with st.expander(f"Ver {len(errores)} errores"):
                                    for i, res in enumerate(errores):
                                        st.error(f"{res['mensaje']}")
                        else:
                            st.info("No se encontraron documentos para procesar")
                        
                        # Ejecutar descargas de manera más simple
                        with st.spinner("Procesando documentos..."):
                            resultados = gestor.ejecutar_descargas()
                            st.session_state.download_results = resultados
                            
                        # Mostrar errores solo si existen
                        errores = [r for r in resultados if not r.get('exito', False)]
                        if errores:
                            with st.expander(f"Ver {len(errores)} errores"):
                                for e in errores:
                                    st.error(f"Error en {e.get('url', 'descarga')}: {e.get('mensaje', 'Error desconocido')}")
                        
                        # Botón para cerrar el diálogo después de completar la descarga
                        if st.button("Cerrar", key="close_dialog"):
                            st.rerun()
                else:
                    st.warning("Debe seleccionar al menos una columna para descargar")
            else:
                st.error("No se encontraron columnas con URLs en este dataframe")
        else:
            st.error("La tabla de datos está vacía. No hay datos para descargar.")
            if st.button("Cerrar", key="close_dialog_empty"):
                st.rerun()
    
    @st.dialog("💾 Descarga de archivos Excel", width="large")
    def configurar_descarga_excel():
        # Selección de fuente de datos
        st.session_state.download_source_excel = st.radio(
            "Seleccione la fuente de datos:",
            options=[
                'df_concentrado', 
                'saved_selections',
                'df_desglosado',
                'saved_selections_desglosado'
            ],
            format_func=lambda x: {
                'df_concentrado': "Tabla Concentrado", 
                'saved_selections': "Tabla Registro temporal (Concentrado)",
                'saved_selections_desglosado': "Tabla Registro temporal (Desglosado)",
                'df_desglosado': "Tabla Desglosado"
            }[x],
            index=0
        )
        
        # Mapeo de opciones a dataframes
        data_source_map = {
            'df_concentrado': st.session_state.df_concentrado,
            'saved_selections': st.session_state.saved_selections,
            'saved_selections_desglosado': st.session_state.saved_selections_desglosado,
            'df_desglosado': st.session_state.df_desglosado
        }
        
        # Verificar si el dataframe seleccionado tiene datos
        selected_df = data_source_map[st.session_state.download_source_excel]
        rows_count = len(selected_df) if not selected_df.empty else 0
        
        st.info(f"📊 Fuente seleccionada: {rows_count:,} registros disponibles")

        if not selected_df.empty:
            # Nombre del archivo
            fecha_actual = time.strftime('%Y-%m-%d')
            nombre_archivo = st.text_input(
                "Nombre del archivo Excel",
                value=f"Reporte_{fecha_actual}",
                help="Nombre del archivo Excel para descarga"
            )
            
            # Botón para descargar el Excel
            if st.button("💾 Generar Excel", key="download_excel_button"):
                try:
                    with st.spinner("🔄 Procesando Excel..."):

                        data_source_map_internal = {
                            'df_concentrado': st.session_state.df_concentrado,
                            'saved_selections': st.session_state.saved_selections,
                            'saved_selections_desglosado': st.session_state.saved_selections_desglosado, 
                            'df_desglosado': st.session_state.df_desglosado
                        }
                        
                        # Crear una copia real para trabajar
                        df_export = data_source_map_internal[st.session_state.download_source_excel].copy()
                        
                        # Identificar tipos de columnas según la selección
                        if st.session_state.download_source_excel in ['df_concentrado', 'saved_selections']:
                            # Formato para datos de contabilidad
                            fecha_cols = ['Fecha Factura', 'Fecha Recepción', 'Fecha Pagado', 'Fecha Autorización']
                            moneda_cols = ['Subtotal', 'Descuento', 'Venta Tasa 0%', 'Venta Tasa 16%', 'IVA 16%', 'ISH', 'Retención IVA', 'Retención ISR', 'Total']
                            numericas_cols = []
                            contabilidad_cols = ['Folio', 'Cuenta Gasto']
                            texto_cols = ['Obra', 'Tipo Gasto', 'Proveedor', 'Residente', 'Estatus', 'Moneda', 'Serie', 'Factura', 'Orden de Compra', 'Remisión', 'UUID']
                        else:
                            # Formato para datos desglosados
                            fecha_cols = ['Fecha Factura', 'Fecha Recepción', 'Fecha Pagado', 'Fecha Autorización']
                            numericas_cols = ['Cantidad']
                            moneda_cols = ['Precio Unitario', 'Subtotal', 'Descuento', 'Venta Tasa 0%', 'Venta Tasa 16%', 'Total IVA', 'Total ISH', 'Retención IVA', 'Retención ISR', 'Total', 'IVA 16%']
                            contabilidad_cols = ['Folio', 'Cuenta Gasto']
                            texto_cols = ['Obra', 'Tipo Gasto', 'Proveedor', 'Residente', 'Estatus', 'Moneda', 'Serie', 'Factura', 'Orden de Compra', 'Remisión', 'UUID']
                        
                        # Convertir columnas especiales manteniendo el tipo de dato correcto
                        for col in df_export.columns:
                            # Convertir fechas a datetime
                            if col in fecha_cols:
                                # Convertir a datetime sin zona horaria
                                df_export[col] = pd.to_datetime(df_export[col], errors='coerce')
                                # Eliminar la información de zona horaria para evitar error en Excel
                                df_export[col] = df_export[col].dt.tz_localize(None) if df_export[col].dt.tz is not None else df_export[col]
                            
                            # Convertir valores monetarios y numéricos a float
                            elif col in moneda_cols or col in numericas_cols:
                                try:
                                    # Asegurar que sea numérico
                                    df_export[col] = pd.to_numeric(df_export[col], errors='coerce')
                                except Exception:
                                    pass
                            
                            # Convertir columnas de contabilidad a enteros
                            elif col in contabilidad_cols:
                                try:
                                    # Asegurar que sea entero
                                    df_export[col] = pd.to_numeric(df_export[col], errors='coerce')
                                    # Convertir a entero donde sea posible
                                    df_export[col] = df_export[col].fillna(0).astype('Int64')  # Int64 permite NaN
                                except Exception:
                                    pass
                        
                        # Guardar en Excel manteniendo formatos
                        output = io.BytesIO()
                        
                        try:
                            # Importaciones necesarias de openpyxl
                            from openpyxl.styles import numbers, Font, PatternFill, Border, Side, Alignment
                            from openpyxl.utils.dataframe import dataframe_to_rows
                            from openpyxl import Workbook
                            
                            # Crear un libro y hoja de trabajo
                            wb = Workbook()
                            ws = wb.active
                            ws.title = "Datos"
                            
                            # Escribir encabezados
                            headers = list(df_export.columns)
                            ws.append(headers)
                            
                            # Definir estilos para encabezados
                            header_font = Font(bold=True, color="FFFFFF")  # Texto en negrita color blanco
                            header_fill = PatternFill(start_color="191970", end_color="191970", fill_type="solid")  # Fondo azul oscuro
                            header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=False)
                            
                            # Definir bordes para todas las celdas
                            thin_border = Border(
                                left=Side(style='thin'), 
                                right=Side(style='thin'), 
                                top=Side(style='thin'), 
                                bottom=Side(style='thin')
                            )
                            
                            # Aplicar formato a los encabezados
                            for col_idx, _ in enumerate(headers, start=1):
                                cell = ws.cell(row=1, column=col_idx)
                                cell.font = header_font
                                cell.fill = header_fill
                                cell.alignment = header_alignment
                                cell.border = thin_border
                            
                            # Escribir datos fila por fila
                            for _, row in df_export.iterrows():
                                # Convertir la fila a lista
                                row_to_write = []
                                for col_name in headers:
                                    cell_value = row[col_name]
                                    row_to_write.append(cell_value)
                                ws.append(row_to_write)
                            
                            # Aplicar formatos a las columnas
                            for col_idx, col_name in enumerate(headers, start=1):
                                # Convertir índice base 1 a letra de columna Excel (A, B, C...)
                                col_letter = ws.cell(row=1, column=col_idx).column_letter
                                
                                # Formato de fechas
                                if col_name in fecha_cols:
                                    for row in range(2, len(df_export) + 2):  # +2 porque Excel empieza en 1 y hay una fila de encabezado
                                        cell = ws.cell(row=row, column=col_idx)
                                        if cell.value is not None and cell.value != '':
                                            cell.number_format = 'DD/MM/YYYY'
                                        # Aplicar bordes
                                        cell.border = thin_border
                                
                                # Formato de moneda
                                elif col_name in moneda_cols:
                                    for row in range(2, len(df_export) + 2):
                                        cell = ws.cell(row=row, column=col_idx)
                                        if cell.value is not None:
                                            cell.number_format = '$#,##0.00'
                                        # Aplicar bordes
                                        cell.border = thin_border
                                
                                # Formato numérico
                                elif col_name in numericas_cols:
                                    for row in range(2, len(df_export) + 2):
                                        cell = ws.cell(row=row, column=col_idx)
                                        if cell.value is not None:
                                            cell.number_format = '#,##0.00'
                                        # Aplicar bordes
                                        cell.border = thin_border
                                        
                                # Formato contabilidad (números enteros sin separador de miles)
                                elif col_name in contabilidad_cols:
                                    for row in range(2, len(df_export) + 2):
                                        cell = ws.cell(row=row, column=col_idx)
                                        if cell.value is not None:
                                            cell.number_format = '0'  # Formato para enteros sin separador
                                        # Aplicar bordes
                                        cell.border = thin_border
                                        
                                # Aplicar solo bordes para las demás columnas
                                else:
                                    for row in range(2, len(df_export) + 2):
                                        cell = ws.cell(row=row, column=col_idx)
                                        cell.border = thin_border
                            
                            # Definir columnas que no deben ser ajustadas en ningún dataframe
                            no_ajustar_general = ['Factura', 'Orden de Compra', 'Remisión', 'UUID', 'sat', 'Folio']
                            
                            # Si estamos en un dataframe desglosado, agregar 'Descripción' a la lista de exclusión
                            no_ajustar = no_ajustar_general.copy()
                            if st.session_state.download_source_excel in ['df_desglosado', 'saved_selections_desglosado']:
                                no_ajustar.append('Descripción')
                            
                            # Ajustar ancho de columnas excepto las excluidas
                            for col in ws.columns:
                                max_length = 0
                                column = col[0].column_letter  # Obtener letra de la columna
                                column_name = ws.cell(row=1, column=col[0].column).value  # Nombre de la columna
                                
                                # Si la columna está en la lista de exclusión, usar ancho predeterminado
                                if column_name in no_ajustar:
                                    continue  # Saltar y dejar el ancho predeterminado
                                
                                # Para las demás columnas, ajustar al contenido
                                for cell in col:
                                    try:
                                        if len(str(cell.value)) > max_length:
                                            max_length = len(str(cell.value))
                                    except:
                                        pass
                                adjusted_width = (max_length + 2)
                                ws.column_dimensions[column].width = adjusted_width
                            
                            # Guardar a BytesIO
                            wb.save(output)
                            
                        except Exception as e:
                            st.error(f"Error generando Excel: {e}")
                            # Mensaje de éxito (sin necesidad de mostrar datos de depuración)
                            raise e
                        
                        # Preparar el archivo para descarga
                        output.seek(0)
                        excel_data = output.getvalue()
                        st.success(f"✅ Archivo Excel listo para descargar")
                        # Generar el enlace de descarga
                        st.download_button(
                            label="📥 Haga clic para descargar",
                            data=excel_data,
                            file_name=f"{nombre_archivo}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )
                    
                    
                except Exception as e:
                    st.error(f"❌ Error al generar el archivo Excel: {e}")
            
            if st.button("Cerrar", key="close_excel_dialog"):
                st.rerun()
        else:
            st.error("La tabla de datos está vacía. No hay datos para descargar.")
            if st.button("Cerrar", key="close_excel_dialog_empty"):
                st.rerun()

    # Botones para abrir los diálogos de descargas
    col_pdf, col_excel = st.sidebar.columns(2)
    if col_pdf.button("📥 Descargar PDFs", key="open_download_dialog"):
        configurar_descarga()
        
    if col_excel.button("📊 Descargar Excel", key="open_excel_download_dialog"):
        configurar_descarga_excel()



        # # Mostrar información sobre las filas actualmente seleccionadas
        # if selection_event.selection and len(selection_event.selection.rows) > 0:
        #     st.info(f"🔍 {len(selection_event.selection.rows)} filas seleccionadas actualmente")



###
#Aqui empezó el código nuevo
####



elif not buscar_button:
    # Si es la carga inicial sin búsqueda activa, mostrar mensaje informativo
    st.info("Utiliza los filtros y haz clic en BUSCAR para ver los datos.")


# Simplificar CSS para evitar conflictos
st.markdown("""
<style>
.simple-header {
    margin-bottom: 1rem;
}
</style>
""", unsafe_allow_html=True)

# # Add help information
# st.markdown("---")
# with st.expander(":information_source: Ayuda para la Base de Datos"):
#     st.markdown("""
#     ### Consejos para usar esta página:
    
#     - **Filtros de datos**: Utiliza los controles de filtrado para buscar información específica.
#     - **Búsqueda**: Usa el campo de búsqueda para encontrar registros por texto.
#     - **Ordenar**: Haz clic en los encabezados de columna para ordenar los datos.
#     - **Exportar**: Puedes seleccionar datos y copiarlos para usarlos en otras aplicaciones.
    
#     Para obtener más ayuda, contacta al administrador del sistema.
#     """)
